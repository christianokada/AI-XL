#written by Scott Bergstresser, Christian Okada, Cameron Mayes, and Brandon Bui
#simple machine learning exercise, takes in categorical data from a training sets
#   and constructs dictionaries that map specific data to specific categories.
#   Program then reads in a second excel file with similar data fields
#   and chooses the best category based on the given data, weighting the
#   categories based on their importance.
#Accenture internship Summer 2017

from __future__ import unicode_literals
from openpyxl import load_workbook


cd = {} # counter dictionary, used to determine the correct category
#dictionaries that map data to the 31 possible categories
ids,office,org,carl,fun,title,tt,tloc,cor,vname,confname,ctype,certname,curl,ccn,act= ({} for i in range(16))
charstr = ['B', 'D', 'F', 'I', 'J', 'L', 'M','P', 'Q', 'S', 'T', 'V', 'W', 'X', 'Z', 'AB']
dlist = []
dlist.extend((ids,office,org,carl,fun,title,tt,tloc,cor,vname,confname,ctype,certname,curl,ccn,act))
KEY = '0xFF001234A117'
TARGET = 'AA'

c = 0
for dic in dlist:
    dic[KEY] = charstr[c]
    c += 1
    
# print dlist

wb = load_workbook('contest-train.xlsx') #crack open a cold workbook with the boys
ws = wb['contest-train']
catlist=[] #populate a list of possible categories
for row in range(2,ws.max_row):
    if ws[TARGET + str(row)].value not in catlist:
        catlist.append(ws[TARGET + str(row)].value)
    if ws[TARGET + str(row)].value not in cd.keys():
        cd[ws[TARGET + str(row)].value] = 0

for cat in catlist:# initialize
    cd[cat] = 0

#initialize dictionaries, these map our data sets to categories. Each dictionary
#has keys representing the 31 categories, values are an array of data learned
#from the training set
for dic in dlist:
    for cat in catlist:
        dic[cat]=[]

#############################Populate the dictionaries####################################


for row in range(2,ws.max_row+1):
    for dic in dlist:
        if (ws[dic.get(KEY)+str(row)].value is not 'NA') and (ws[dic.get(KEY)+str(row)].value is not 'N/A'):
            if ws[dic.get(KEY)+str(row)].value not in ids[ws[TARGET+str(row)].value]:
                dic[ws[TARGET+str(row)].value].append(ws[dic.get(KEY)+str(row)].value)#.encode("ascii", "ignore"))

#################Dictionaries have been created, now to parse test set###################

wb2 = load_workbook('contest-test.xlsx') #cold workbook number 2
ws2 = wb2['contest-test']

#Data have different weights based on importance to categories from training set

for row in range(2,ws2.max_row):
    for cat in catlist:
        if ws2['D'+str(row)].value in office[cat]:
            cd[cat] += 1
        if ws2['F'+str(row)].value in org[cat]:
            cd[cat] += 2
        if ws2['I'+str(row)].value in carl[cat]:
            cd[cat] += 1
        if ws2['J'+str(row)].value in fun[cat]:
            cd[cat] += 1
        if ws2['L'+str(row)].value in title[cat]:
            cd[cat] += 5
        if ws2['M'+str(row)].value in tt[cat]:
            cd[cat] += 3
        if ws2['P'+str(row)].value in tloc[cat]:
            cd[cat] += 4
        if ws2['Q'+str(row)].value in cor[cat]:
            cd[cat] += 4
        if ws2['S'+str(row)].value in vname[cat]:
            cd[cat] += 2
        if ws2['T'+str(row)].value in confname[cat]:
            cd[cat] += 2
        if ws2['V'+str(row)].value in ctype[cat]:
            cd[cat] += 2
        if ws2['W'+str(row)].value in certname[cat]:
            cd[cat] += 3
        if ws2['X'+str(row)].value in curl[cat]:
            cd[cat] += 3
        if ws2['Z'+str(row)].value in ccn[cat]:
            cd[cat] += 2
        if ws2['AB'+str(row)].value in act[cat]:
            cd[cat] += 6
        
    maximum = max(cd, key=cd.get)
    maxcount = cd[maximum]
    ws2[TARGET+str(row)].value=maximum
    for key in cd.keys():
        cd[key] = 0

wb2.save('contest-test.xlsx')
